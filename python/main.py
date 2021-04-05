import os
import requests
import time

uid = os.environ['UID42']
secret = os.environ['SECRET42']
url = 'https://api.intra.42.fr/'
urlv2 = 'https://api.intra.42.fr/v2/'

def getToken():
  payload = {
    'grant_type': 'client_credentials',
    'client_id': uid,
    'client_secret': secret
  }
  r = requests.post(url + 'oauth/token', data=payload)
  return r.json()['access_token']

def prepareApiCalls():
  global token, headers
  token = getToken()
  headers = {'Authorization': 'Bearer {}'.format(token)}

def getFromApi(path, params={}):
  r = requests.get(urlv2 + path, headers=headers, params=params)
  if r.status_code != requests.codes.ok:
    return (-1)
  else:
    return(r.json())

def getUser(login):
  return getFromApi('users/'+login)

objet = getUser('nklarsfe')
userName = objet['login']
projectName = objet['projects_users'][0]['project']['name']
wasValidated = objet['projects_users'][0]['validated?']
import arrow
text = objet['projects_users'][0]['marked_at']
date = arrow.get(text)
dateAsFloat = float(date.format('X'))
dateInDays = (dateAsFloat / (3600*24)) - 18175

studsNames = []


# ----------------------------------
# pickle

import pickle
def save_object(obj, filename):
  with open(filename, 'wb') as output:  # Overwrites any existing file.
    pickle.dump(obj, output, pickle.HIGHEST_PROTOCOL)

def get_object(filename):
  with open(filename, 'rb') as input:
    return pickle.load(input)

# toto={'a':1, 'b':2}
# save_object(studs, '../dataFrom42api/studsOld.pkl')
# studs = get_object('../dataFrom42api/studs.pkl')

# -------------------------------------------------

def getNamesInPool(year, month):
  global names
  prepareApiCalls()
  params = {
    'filter[pool_year]': year,
    'filter[pool_month]': month,
    'filter[primary_campus_id]': '1',
    'page[number]': '0',
    'page[size]': '100'
  }
  while True:
    pageNumber = int(params['page[number]'])
    print(pageNumber)
    params['page[number]'] = str(pageNumber + 1)
    res = -1
    while res == -1:
      print('...trying...')
      res = getFromApi('users', params)
      # if (res == -1):
      #   prepareApiCalls()
    for usr in res:
      names.append(usr['login'])
    if len(res) != 100:
      break

names = []
for year in range(2013, 2019):
  print(year)
  for month in ['july', 'august', 'september']:
    print(month)
    getNamesInPool(year, month)

myIterator = filter(lambda stud: stud[0] != '3', names)
nonAnonymizedNames = list(myIterator)
names = nonAnonymizedNames

names = get_object('names.pkl')

# save_object(names, '../dataFrom42api/names.pkl')
# matches = (x for x in names if x[0] == 'c' and x[1] == 'a')
# list(matches)

# -----------------------------------------------
# on recupere les donnees dans studs

# studs = get_object('../dataFrom42api/old_studs.pkl')

# studs = {}
# compteur = 0 
# for i in range(0,len(names)): 
#   gotIt = False
#   while not gotIt:
#     objet = getUser(names[i])
#     gotIt = (objet != -1)
#     if not gotIt:
#       compteur += 1
#       print(i, 'fail', compteur)
#       if compteur >= 10:
#         time.sleep(float(1))
#         prepareApiCalls()
#         compteur = 0
#         time.sleep(float(1))
#   studs[names[i]] = objet['projects_users']
#   print(i, 'ok', names[i])

# save_object(studs, '../dataFrom42api/studsOld.pkl')

# ------------------------------------
# on mets les donnees dans un format avec juste l'essentiel

import arrow

# studs2={}
# for name, projects in studs.items():
#   studs2[name] = {}
#   for project in projects:
#     if len(project['cursus_ids']) >= 1:
#       if project['cursus_ids'][0] == 1: # 21
#         nameProject = project['project']['name']
#         wasValidated = project['validated?']
#         text = project['marked_at']
#         if text != None:
#           date = arrow.get(text)
#           dateAsFloat = float(date.format('X'))
#           dateInDays = (dateAsFloat / (3600*24)) - 16028 # 18175
#           studs2[name][nameProject] = {
#             'wasValidated':wasValidated,
#             'dateInDays':dateInDays
#           }

#--------------------------
# 2013: 18 novembre 2013
# 2014: 20 novembre 2014 ?
# 2015: 26 novembre 2015 ?
# 2016: 26 septembre 2016 ?
# 2017: 29 septembre 2017 ?
# 2018: 24 septembre 2018 ?

# debut=[388, 813, 1255, 1888, 2494, 3334]
# decalage = [367, 738, 1043, 1411, 1771]
# for i in range(len(debut) - 1):
#   for j in range(debut[i], debut[i+1]):
#     projects = studs2[names[j]]
#     for nameProject, details in projects.items():
#       details['dateInDays'] -= decalage[i]

# save_object(studs2, '../dataFrom42api/old_studs2_2.pkl')
# studs2 = get_object('../dataFrom42api/old_studs2_2.pkl')

# ----------------------------
# on cree le fichier excel

# projectsNames = [
#   'Libft', 
#   'netwhat', 'get_next_line', 'ft_printf', 
#   'Exam Rank 02', 'ft_server', 'miniRT', 'cub3d',
#   'Exam Rank 03', 'libasm', 'ft_services', 'minishell',
#   'Exam Rank 04', 'CPP Module 08', 'Philosophers',
#   'Exam Rank 05', 'ft_containers', 'ft_irc', 'webserv',
#   'Exam Rank 06', 'ft_transcendence'
# ]

# Piscine reloaded and Fillit starting 2016 (2013 2014 and 2015 without them)
projectNames_0 = ['Piscine Reloaded', 'Libft','GET_Next_Line', 'Fillit'] 
# (Wolf3d + Doom) or (RTv1 + RT)
projectNames_graphics=['FdF', "Fract'ol", 'Wolf3d', 'Doom Nukem', 'RTv1', 'RT'] # GUImp
projectNames_system=['ft_ls', 'minishell', '21sh', '42sh'] # ft_select
# Push_swap or Filler
projectNames_security=['ft_printf', 'Push_swap', 'Filler', 'Lem_in', 'Corewar'] # mod1
projectNames_web=['Piscine PHP','Camagru', 'Matcha', 'Hypertube']
internship = ['Contract Upload', 'Company mid evaluation', 'Company final evaluation', 'First Internship']

projectsNames = projectNames_0 + projectNames_graphics + projectNames_system + projectNames_security + projectNames_web + internship

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

for i in range(len(projectsNames)):
  ws.cell(row=1, column=i+2).value = projectsNames[i]

# name='nklarsfe'
# projects = studs2[name]
indexName = 2
# for name, projects in studs2.items():
for i in range(0, len(names)):
  name = names[i]
  projects = studs2[name]
  ws.cell(row=indexName, column=1).value = name
  for i in range(len(projectsNames)):
    projectName = projectsNames[i]
    if(projectName in projects):
      # print(i, projectName, projects[projectName])
      if(projects[projectName]['wasValidated']):
        dateInDays = projects[projectName]['dateInDays']
        ws.cell(row=indexName, column=i+2).value = dateInDays
  indexName += 1
  print(indexName)

wb.save('old_result4.xlsx')

# for key, value in studs.items():

